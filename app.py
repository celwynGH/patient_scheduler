# app.py
"""
Patient Scheduler Web App (Excel version) with Search
-----------------------------------------------------
- Add, delete, list appointments
- Search appointments by name
- Auto save/load from appointments.xlsx
- Limit patients per hour
- Download Excel file directly

Requirements:
  pip install flask openpyxl
"""

from flask import Flask, request, jsonify, send_file, render_template_string
from openpyxl import Workbook, load_workbook
import os, uuid, tempfile
from datetime import datetime

# ---------- CONFIG ----------
EXCEL_FILE = "appointments.xlsx"
MAX_PER_HOUR = 4
DATETIME_FORMAT = "%Y-%m-%dT%H:%M"
# ----------------------------

app = Flask(__name__)
appointments = []
HEADERS = ["id", "name", "address", "reason", "datetime", "created_at"]

# ---------- EXCEL HELPERS ----------
def load_excel():
    global appointments
    appointments = []
    if not os.path.exists(EXCEL_FILE):
        return
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        appt = dict(zip(HEADERS, row))
        appointments.append(appt)

def save_excel():
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for a in appointments:
        ws.append([a[h] for h in HEADERS])
    tmp_fd, tmp_path = tempfile.mkstemp(prefix="appts-", suffix=".xlsx")
    os.close(tmp_fd)
    wb.save(tmp_path)
    os.replace(tmp_path, EXCEL_FILE)

def count_in_hour(dt_iso):
    try:
        dt = datetime.fromisoformat(dt_iso)
    except Exception:
        return 0
    count = 0
    for a in appointments:
        try:
            adt = datetime.fromisoformat(a["datetime"])
            if (adt.year, adt.month, adt.day, adt.hour) == (dt.year, dt.month, dt.day, dt.hour):
                count += 1
        except Exception:
            continue
    return count
# -----------------------------------

@app.route("/")
def index():
    return render_template_string(INDEX_HTML, max_per_hour=MAX_PER_HOUR)

@app.route("/api/appointments", methods=["GET", "POST", "DELETE"])
def api_appointments():
    if request.method == "GET":
        return jsonify({"ok": True, "appointments": sorted(appointments, key=lambda a: a["datetime"])})

    if request.method == "POST":
        data = request.json or {}
        name = data.get("name", "").strip()
        address = data.get("address", "").strip()
        reason = data.get("reason", "").strip()
        dt = data.get("datetime", "").strip()

        if not name or not dt:
            return jsonify({"ok": False, "error": "Name and datetime are required."}), 400

        try:
            datetime.fromisoformat(dt)
        except Exception:
            return jsonify({"ok": False, "error": "Invalid datetime format."}), 400

        if count_in_hour(dt) >= MAX_PER_HOUR:
            return jsonify({"ok": False, "error": f"Hour full (max {MAX_PER_HOUR})."}), 409

        new_appt = {
            "id": str(uuid.uuid4()),
            "name": name,
            "address": address,
            "reason": reason,
            "datetime": dt,
            "created_at": datetime.utcnow().isoformat()
        }
        appointments.append(new_appt)
        save_excel()
        return jsonify({"ok": True, "appointment": new_appt}), 201

    if request.method == "DELETE":
        data = request.json or {}
        appt_id = data.get("id")
        before = len(appointments)
        appointments[:] = [a for a in appointments if a["id"] != appt_id]
        if len(appointments) == before:
            return jsonify({"ok": False, "error": "Appointment not found"}), 404
        save_excel()
        return jsonify({"ok": True})

@app.route("/download")
def download_excel():
    save_excel()
    return send_file(EXCEL_FILE, as_attachment=True, download_name=EXCEL_FILE)

# Load at startup
load_excel()

# ---------- HTML (UI with Search) ----------
INDEX_HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <title>Patient Scheduler</title>
  <link rel="icon" type="image/png" href="/static/hospital.png">
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <style>
    body { font-family: system-ui, -apple-system, "Segoe UI", Roboto, "Helvetica Neue", Arial; margin: 20px; }
    h1 { margin-bottom: 6px; }
    form { border: 1px solid #ddd; padding: 12px; border-radius: 8px; max-width: 720px; }
    label { display:block; margin-top:8px; font-weight:600; }
    input[type="text"], input[type="datetime-local"], textarea { width:100%; padding:8px; border-radius:6px; border:1px solid #ccc; }
    button { margin-top:10px; padding:8px 12px; border-radius:6px; cursor:pointer; }
    table { width:100%; border-collapse: collapse; margin-top:18px; }
    th, td { padding:8px; border-bottom:1px solid #eee; text-align:left; }
    .small { font-size:0.9rem; color:#666; }
    .row { display:flex; gap:12px; align-items:center; }
    .col { flex:1; }
    .actions { white-space:nowrap; }
    .error { color: #a00; margin-top:8px; }
    .success { color: #080; margin-top:8px; }
    .topbar { display:flex; gap:12px; align-items:center; margin-bottom:12px; }
    .limit-badge { background:#f0f4ff; padding:6px 8px; border-radius:6px; border:1px solid #c7d2ff; }
  </style>
</head>
<body>
  <div class="topbar">
    <img src="/static/hospital.png" alt="Logo" style="height:40px; margin-right:12px;">
    <h1>Patient Scheduler</h1>
    <div class="limit-badge small">Limit per hour: <strong>{{ max_per_hour }}</strong></div>
    <div style="margin-left:auto;">
      <a href="/download"><button>Download Excel File</button></a>
    </div>
  </div>

  <form id="apptForm" onsubmit="return false;">
    <label>Complete Name *</label>
    <input id="name" type="text" placeholder="John Doe" required>

    <label>Address</label>
    <input id="address" type="text" placeholder="123 Street, City">

    <label>Reason for check-up</label>
    <textarea id="reason" rows="2" placeholder="e.g. Regular check-up, cough"></textarea>

    <label>Schedule Date & Time *</label>
    <input id="datetime" type="datetime-local" required>

    <div style="display:flex; gap:10px;">
      <button id="addBtn" onclick="addAppointment()">Add appointment</button>
      <button type="button" onclick="clearForm()">Clear</button>
      <div id="message" style="margin-left:12px;"></div>
    </div>
  </form>

  <!-- SEARCH BOX -->
  <div style="margin-top:12px; margin-bottom:6px;">
    <label>Search by Name:</label>
    <input id="searchName" type="text" placeholder="Enter name to search..." oninput="renderTable()">
  </div>

  <h2 style="margin-top:24px;">All Appointments</h2>
  <div id="tableWrap"></div>

<script>
const apiBase = "/api/appointments";

async function fetchAppointments() {
  const res = await fetch(apiBase);
  const data = await res.json();
  if (!data.ok) { console.error("Failed to load"); return []; }
  return data.appointments || [];
}

function formatLocal(dtIso) {
  try {
    const d = new Date(dtIso);
    if (isNaN(d)) return dtIso;
    return d.toLocaleString();
  } catch (e) {
    return dtIso;
  }
}

async function renderTable() {
  const appts = await fetchAppointments();
  const searchValue = document.getElementById("searchName").value.trim().toLowerCase();

  const filtered = appts.filter(a => a.name.toLowerCase().includes(searchValue));

  if (!filtered.length) {
    document.getElementById("tableWrap").innerHTML = "<p class='small'>No appointments found.</p>";
    return;
  }

  let html = "<table><thead><tr><th>Name</th><th>Address</th><th>Reason</th><th>Schedule</th><th class='actions'>Actions</th></tr></thead><tbody>";
  for (const a of filtered) {
    html += `<tr>
      <td>${escapeHtml(a.name)}</td>
      <td>${escapeHtml(a.address)}</td>
      <td>${escapeHtml(a.reason)}</td>
      <td>${escapeHtml(a.datetime)}</td>
      <td class="actions">
        <button onclick="deleteAppointment('${a.id}')">Delete</button>
      </td>
    </tr>`;
  }
  html += "</tbody></table>";
  document.getElementById("tableWrap").innerHTML = html;
}

function escapeHtml(s) {
  if (!s) return "";
  return s.replaceAll("&", "&amp;").replaceAll("<", "&lt;").replaceAll(">", "&gt;");
}

function showMessage(msg, ok=true) {
  const el = document.getElementById("message");
  el.innerText = msg;
  el.className = ok ? "success" : "error";
  setTimeout(()=>{ el.innerText = ""; el.className = ""; }, 5000);
}

function clearForm() {
  document.getElementById("name").value = "";
  document.getElementById("address").value = "";
  document.getElementById("reason").value = "";
  document.getElementById("datetime").value = "";
}

async function addAppointment() {
  const name = document.getElementById("name").value.trim();
  const address = document.getElementById("address").value.trim();
  const reason = document.getElementById("reason").value.trim();
  const datetime = document.getElementById("datetime").value;

  if (!name || !datetime) {
    showMessage("Name and datetime are required.", false);
    return;
  }

  const payload = { name, address, reason, datetime };
  const res = await fetch(apiBase, {
    method: "POST",
    headers: {"Content-Type":"application/json"},
    body: JSON.stringify(payload)
  });

  if (res.status === 201) {
    showMessage("Appointment added.");
    clearForm();
    renderTable();
  } else {
    const json = await res.json().catch(()=>({error:"Server error"}));
    showMessage(json.error || "Failed to add.", false);
  }
}

async function deleteAppointment(id) {
  if (!confirm("Delete this appointment?")) return;
  const res = await fetch(apiBase, {
    method: "DELETE",
    headers: {"Content-Type":"application/json"},
    body: JSON.stringify({id})
  });
  const json = await res.json();
  if (json.ok) {
    showMessage("Deleted.");
    renderTable();
  } else {
    showMessage(json.error || "Failed to delete", false);
  }
}

document.addEventListener("DOMContentLoaded", () => {
  renderTable();
});
</script>
</body>
</html>
"""

if __name__ == "__main__":
    app.run(debug=True)
